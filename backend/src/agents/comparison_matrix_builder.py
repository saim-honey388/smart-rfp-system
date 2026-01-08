"""
Comparison Matrix Builder

Combines the blank RFP structure with multiple vendor proposals
to create a multi-vendor comparison matrix.
"""

from typing import List, Dict, Any, Optional
from pydantic import BaseModel, Field
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import difflib

from backend.src.agents.form_structure_analyzer import (
    ProposalFormStructure,
    DiscoveredFormRow
)
from backend.src.agents.vendor_data_extractor import (
    VendorProposalData,
    FilledFormRow
)


class ComparisonMatrixBuilder:
    """
    Builds multi-vendor comparison matrices from RFP structure and vendor data.
    
    Output: Excel file matching the format of AV - Bid Analysis Spreadsheet.xlsx
    """
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_font = Font(bold=True)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def build_comparison_dataframe(
        self,
        rfp_structure: Optional[ProposalFormStructure],
        vendor_proposals: List[VendorProposalData]
    ) -> pd.DataFrame:
        """
        Build a pandas DataFrame for the comparison matrix.
        Dynamically adapts to the columns defined in rfp_structure.
        """
        # If no RFP structure, elect one from the proposals (Consensus Logic)
        if not rfp_structure or not rfp_structure.rows:
            print("⚠ Missing RFP structure - Electing consensus structure from proposals...")
            rfp_structure = self._elect_structure_from_proposals(vendor_proposals)
            if not rfp_structure:
                raise ValueError("Cannot build comparison: No RFP structure and no valid vendor proposals to infer from.")

        rows = []
        
        # Create vendor lookup by item_id for each proposal
        vendor_lookups = {}
        for proposal in vendor_proposals:
            vendor_lookups[proposal.vendor_name] = {
                row.item_id: row for row in proposal.filled_rows
            }
        
        # Track current section for section headers
        current_section = None
        
        # Determine columns to display
        fixed_cols = rfp_structure.fixed_columns or ["Item", "Description"]
        vendor_cols = rfp_structure.vendor_columns or ["Quantity", "Unit", "Unit Cost", "Total"]
        
        for i, rfp_row in enumerate(rfp_structure.rows):
            row_data = {}
            
            # Add section header row if section changed
            if rfp_row.section and rfp_row.section != current_section:
                current_section = rfp_row.section
                section_row = {fixed_cols[0]: rfp_row.section}  # Put section name in first column
                # Fill rest with empty
                for col in fixed_cols[1:]:
                    section_row[col] = ""
                for proposal in vendor_proposals:
                    for v_col in vendor_cols:
                        section_row[f"{proposal.vendor_name} {v_col}"] = ""
                rows.append(section_row)
            
            # Fixed columns (from RFP)
            # Map known fixed columns or use defaults
            # Try to get Item ID from attribute or values
            item_val = rfp_row.item_id
            if not item_val:
                item_val = self._get_value_insensitive(rfp_row.values, fixed_cols[0])
            row_data[fixed_cols[0]] = item_val

            if len(fixed_cols) > 1:
                # Try to get Description from attribute or values
                desc_val = rfp_row.description
                if not desc_val:
                    desc_val = self._get_value_insensitive(rfp_row.values, fixed_cols[1]) # e.g. "Description"
                row_data[fixed_cols[1]] = desc_val
            
            # Additional fixed columns if any (from values dict if available)
            # (Simplification: assuming mostly Item/Desc for fixed)
            
                # Strategy:
                # 1. Exact ID Match (already tried via vendor_lookups)
                # 2. Semantic/Fuzzy Match on Description
                # 3. Positional Fallback (Index)
                
                vendor_row = vendor_lookups[proposal.vendor_name].get(rfp_row.item_id)
                
                if not vendor_row:
                    # Semantic search
                    vendor_row = self._find_best_match_row(rfp_row, proposal.filled_rows)
                
                if not vendor_row and i < len(proposal.filled_rows):
                    # Final Fallback: Index Alignment
                    vendor_row = proposal.filled_rows[i]
                
                for v_col in vendor_cols:
                    col_key = f"{proposal.vendor_name} {v_col}"
                    
                    if vendor_row:
                        # Try to get value from vendor's values dict (case-insensitive lookup preferred)
                        val = self._get_value_insensitive(vendor_row.values, v_col)
                        
                        # Fallback: if value missing in vendor, try to get from RFP (e.g. Quantity provided by RFP)
                        if not val:
                            val = self._get_value_insensitive(rfp_row.values, v_col)
                            
                        row_data[col_key] = val or ""
                    else:
                        # No vendor data for this item -> check if RFP has a value (e.g. Qty)
                        val = self._get_value_insensitive(rfp_row.values, v_col)
                        row_data[col_key] = val or ("-" if "Cost" in v_col or "Total" in v_col else "")

            
            rows.append(row_data)
        
        # Add Grand Total row
        total_row = {fixed_cols[0]: "", fixed_cols[1] if len(fixed_cols) > 1 else fixed_cols[0]: "GRAND TOTAL"}
        for proposal in vendor_proposals:
            # Add empty cells for non-total columns
            for v_col in vendor_cols:
                col_key = f"{proposal.vendor_name} {v_col}"
                if "Total" in v_col or "Amount" in v_col or "Cost" in v_col:
                     # Put grand total in the last column usually? 
                     # For now, put it in the column that sounds like 'Total'
                     if v_col.lower() in ["total", "total amount", "total cost", "amount"]:
                         total_row[col_key] = proposal.grand_total or ""
                     else:
                         total_row[col_key] = ""
                else:
                    total_row[col_key] = ""
                    
        rows.append(total_row)
        
        return pd.DataFrame(rows)

    def _get_value_insensitive(self, values_dict: Dict[str, Any], key: str) -> str:
        """Helper to get value from dict with case-insensitive key."""
        if not values_dict:
            return ""
        # Direct match
        if key in values_dict:
            return str(values_dict[key])
        # Case insensitive
        key_lower = key.lower()
        for k, v in values_dict.items():
            if k.lower() == key_lower:
                return str(v)
        return ""

    def _elect_structure_from_proposals(self, vendor_proposals: List[VendorProposalData]) -> Optional[ProposalFormStructure]:
        """
        Consensus Logic: Elects a surrogate RFP structure from the submitted proposals.
        Strategy: Use the proposal with the most common row count (majority rule).
        """
        if not vendor_proposals:
            return None
            
        # Group by row count
        from collections import Counter
        counts = Counter([len(p.filled_rows) for p in vendor_proposals])
        
        if not counts:
            return None
            
        # Find the most frequent row count(s)
        most_common = counts.most_common()
        max_freq = most_common[0][1]
        
        # Get all candidates that share the max frequency (e.g. if 20 and 25 both appear once)
        candidates = [count for count, freq in most_common if freq == max_freq]
        
        # Tie-breaker: Pick the largest row count to avoid hiding data
        winner_count = max(candidates)
        
        print(f"  Election: Counts={dict(counts)}. Candidates with freq {max_freq}: {candidates}. Winner={winner_count}")
        
        # Find the first proposal that matches this winner count
        winner = next(p for p in vendor_proposals if len(p.filled_rows) == winner_count)
        print(f"  Election: Winner is {winner.vendor_name} with {len(winner.filled_rows)} rows.")
        
        # Infer columns from the first row of valid data
        fixed_cols = []
        if winner.filled_rows:
            sample = winner.filled_rows[0]
            # Only include Item/Description if they have data or are in values
            if sample.item_id or self._get_value_insensitive(sample.values, "Item"):
                fixed_cols.append("Item")
            if sample.description or self._get_value_insensitive(sample.values, "Description"):
                fixed_cols.append("Description")
        
        # Ensure at least one fixed col for tracking/indenting?
        if not fixed_cols:
            fixed_cols = ["Item"]
        
        # KEY CHANGE: Dynamically fetch columns from the winner's data
        # Check the first row's values to see what columns exist
        vendor_cols = []
        if winner.filled_rows:
            # Get all keys from the values dict (which holds the dynamic columns)
            # Use the first row as a sample
            sample_values = winner.filled_rows[0].values
            if sample_values:
                vendor_cols = list(sample_values.keys())
                print(f"  Election: Discovered dynamic columns from DB: {vendor_cols}")
        
        # Fallback if no dynamic columns found (e.g. valid structure but empty values?)
        if not vendor_cols:
             vendor_cols = ["Quantity", "Unit", "Unit Cost", "Total"]
             print("  Election: Warning - No dynamic columns found in values. Using defaults.")
        

        # Convert FilledFormRows to DiscoveredFormRows
        new_rows = []
        for row in winner.filled_rows:
            new_rows.append(DiscoveredFormRow(
                section=row.section,
                item_id=row.item_id,
                description=row.description,
                values=row.values # Reuse the values (Qty, Unit, etc) as defaults
            ))
            
        return ProposalFormStructure(
            form_title=f"Consensus Structure (from {winner.vendor_name})",
            tables=[],
            fixed_columns=fixed_cols,
            vendor_columns=vendor_cols,
            sections=sorted(list(set(r.section for r in new_rows if r.section))),
            rows=new_rows
        )
    
    def build_comparison_excel(
        self,
        rfp_structure: Optional[ProposalFormStructure],
        vendor_proposals: List[VendorProposalData],
        output_path: str,
        include_vendor_info: bool = True
    ) -> str:
        """
        Build a formatted Excel comparison matrix.
        """
        print(f"--- Building Comparison Matrix for {len(vendor_proposals)} vendors ---")
        
        # Ensure we have a structure (Consensus Logic)
        if not rfp_structure or not rfp_structure.rows:
            print("⚠ (Excel) Missing RFP structure - Electing consensus structure...")
            rfp_structure = self._elect_structure_from_proposals(vendor_proposals)
            if not rfp_structure:
                 raise ValueError("Cannot build Excel: No RFP structure and no valid vendor proposals.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bid Comparison"
        
        current_row = 1
        
        # Add RFP title
        ws.cell(row=current_row, column=1, value=rfp_structure.form_title)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2
        
        # Dynamic Column Counts
        fixed_count = len(rfp_structure.fixed_columns or ["Item", "Description"])
        vendor_col_count = len(rfp_structure.vendor_columns or ["Quantity", "Unit", "Unit Cost", "Total"])
        
        if include_vendor_info:
            # Add vendor info header
            col = fixed_count + 1  # Start after fixed columns
            for proposal in vendor_proposals:
                ws.cell(row=current_row, column=col, value=proposal.vendor_name)
                ws.cell(row=current_row, column=col).font = Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + vendor_col_count - 1)
                
                if proposal.vendor_contact:
                    ws.cell(row=current_row + 1, column=col, value=proposal.vendor_contact)
                if proposal.vendor_license:
                    ws.cell(row=current_row + 2, column=col, value=f"License: {proposal.vendor_license}")
                
                col += vendor_col_count  # Move to next vendor block
            current_row += 4
        
        # Build DataFrame (passing the elected structure)
        df = self.build_comparison_dataframe(rfp_structure, vendor_proposals)
        
        # Add column headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.thin_border
        current_row += 1
        
        # Add data rows
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = self.thin_border
                
                # Format section headers
                if col_idx == 1 and str(value).startswith(('I ', 'II ', 'III ', 'IV ', 'V ')):
                    cell.font = self.section_font
                
                # Format currency values
                if isinstance(value, str) and value.startswith('$'):
                    cell.alignment = Alignment(horizontal='right')
                    
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8    # Item
        ws.column_dimensions['B'].width = 50   # Description
        
        for col in range(3, len(df.columns) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 12
        
        # Save
        wb.save(output_path)
        print(f"  ✓ Saved comparison matrix to: {output_path}")
        
        return output_path
    
    def _find_best_match_row(self, rfp_row: DiscoveredFormRow, candidate_rows: List[FilledFormRow]) -> Optional[FilledFormRow]:
        """
        Find best matching row in candidates using fuzzy semantic matching on Description.
        Used when exact Item ID match fails.
        """
        if not candidate_rows:
            return None
            
        best_row = None
        best_score = 0.0
        
        # Target description (from RFP or Elected Structure)
        target_desc = rfp_row.description or self._get_value_insensitive(rfp_row.values, "Description") or ""
        if not target_desc:
            return None # Can't match on empty description
            
        for row in candidate_rows:
            # Candidate description
            cand_desc = row.description or self._get_value_insensitive(row.values, "Description") or ""
            if not cand_desc:
                continue
                
            # Calculate similarity
            score = difflib.SequenceMatcher(None, target_desc.lower(), cand_desc.lower()).ratio()
            
            if score > best_score:
                best_score = score
                best_row = row
        
        # Threshold: 0.6 means reasonable similarity (e.g. "Roof Repair" vs "Roof Repairs")
        # If very different ("Alpha" vs "Qwertt"), this will fail, and we fall back to Index matching.
        if best_score > 0.6:
            return best_row
            
        return None

    def build_from_selected_proposals(
        self,
        rfp_structure: Optional[ProposalFormStructure],
        all_proposals: List[VendorProposalData],
        selected_proposal_ids: List[str],
        output_path: str
    ) -> str:
        """
        Build comparison matrix for only selected/accepted proposals.
        
        Args:
            rfp_structure: RFP form structure
            all_proposals: All vendor proposals (saved in DB)
            selected_proposal_ids: IDs of proposals to include in comparison
            output_path: Output file path
        """
        # Filter to selected proposals only
        selected_proposals = [
            p for p in all_proposals 
            if p.proposal_id in selected_proposal_ids
        ]
        
        if not selected_proposals:
            raise ValueError("No matching proposals found for the selected IDs")
        
        print(f"  Building comparison for {len(selected_proposals)}/{len(all_proposals)} proposals")
        
        return self.build_comparison_excel(
            rfp_structure,
            selected_proposals,
            output_path
        )


# --- Convenience Functions ---

def generate_comparison_report(
    rfp_structure: ProposalFormStructure,
    vendor_proposals: List[VendorProposalData],
    output_dir: str = ".",
    filename: str = "Comparison_Matrix.xlsx"
) -> str:
    """
    Convenience function to generate a comparison report.
    """
    builder = ComparisonMatrixBuilder()
    output_path = os.path.join(output_dir, filename)
    return builder.build_comparison_excel(rfp_structure, vendor_proposals, output_path)


# --- Test ---
if __name__ == "__main__":
    print("=== Testing Comparison Matrix Builder ===\n")
    
    # Create mock data for testing
    from backend.src.agents.form_structure_analyzer import DiscoveredFormRow
    
    # Mock RFP structure
    mock_structure = ProposalFormStructure(
        form_title="AUDUBON VILLAS CONDOMINIUM - REPAIR SPECIFICATIONS",
        tables=[],
        fixed_columns=["Item", "Description of Work"],
        vendor_columns=["Quantity", "Unit", "Unit Cost", "Total"],
        sections=["I Structural", "II Balcony Restoration"],
        rows=[
            DiscoveredFormRow(section="I Structural", item_id="1", description="Wall sheathing repairs and replacement as needed.", values={"quantity": "TBD", "unit": "SF"}),
            DiscoveredFormRow(section="I Structural", item_id="2", description="Wall framing members repairs and replacement as needed.", values={"quantity": "TBD", "unit": "LF"}),
            DiscoveredFormRow(section="II Balcony Restoration", item_id="1", description="Remove and replace all existing ceiling finishes.", values={"quantity": "13,450", "unit": "SF"}),
        ]
    )
    
    # Mock vendor data
    mock_vendors = [
        VendorProposalData(
            proposal_id="prop-001",
            rfp_id="rfp-audubon",
            vendor_name="DueAll",
            filled_rows=[
                FilledFormRow(section="I Structural", item_id="1", description="Wall sheathing repairs", quantity="TBD", unit="SF", unit_cost="$4.10", total="$4.10"),
                FilledFormRow(section="I Structural", item_id="2", description="Wall framing members", quantity="TBD", unit="LF", unit_cost="$7.49", total="$7.49"),
                FilledFormRow(section="II Balcony Restoration", item_id="1", description="Remove and replace ceiling", quantity="13,450", unit="SF", unit_cost="$9.75", total="$131,137.50"),
            ],
            grand_total="$1,122,772.91"
        ),
        VendorProposalData(
            proposal_id="prop-002",
            rfp_id="rfp-audubon",
            vendor_name="IECON",
            filled_rows=[
                FilledFormRow(section="I Structural", item_id="1", description="Wall sheathing repairs", quantity="TBD", unit="SF", unit_cost="$8.00", total="$0"),
                FilledFormRow(section="I Structural", item_id="2", description="Wall framing members", quantity="TBD", unit="LF", unit_cost="$15.00", total="$0"),
                FilledFormRow(section="II Balcony Restoration", item_id="1", description="Remove and replace ceiling", quantity="13,450", unit="SF", unit_cost="$6.50", total="$87,425.00"),
            ],
            grand_total="$989,500.00"
        )
    ]
    
    # Generate comparison
    builder = ComparisonMatrixBuilder()
    output_path = builder.build_comparison_excel(
        mock_structure,
        mock_vendors,
        "Test_Comparison_Matrix.xlsx"
    )
    
    print(f"\nGenerated: {output_path}")
